"""
Importador de inventario inicial (#2 parcial): valida el .xlsx fila por fila y,
si no hay errores, crea los productos con su stock inicial en el kardex (atómico).
"""
import io
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.modules.inventario.importador import (
    EXPECTED_HEADERS, validar, importar, generar_plantilla,
)
from app.modules.ventas.models import Producto
from app.modules.inventario.models import MovimientoInventario
from app.modules.usuarios.models import Usuario


def _xlsx(rows, headers=EXPECTED_HEADERS, hoja="Inventario") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(list(headers))
    for r in rows:
        ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FILA_OK = {
    "sku": "IMP-001", "nombre": "Producto Uno", "categoria": "Biocida",
    "marca": "Superozono", "unidad_medida": "Litro",
    "precio_venta": 85000, "tarifa_iva": 19, "stock_actual": 120,
    "precio_costo": 32000, "stock_minimo": 20,
}


async def _admin(db):
    return (await db.execute(
        select(Usuario).where(Usuario.email == "admin@test.com"))).scalar_one()


@pytest.mark.asyncio
async def test_valida_y_importa_ok(db_session):
    fila2 = {**FILA_OK, "sku": "IMP-002", "nombre": "Producto Dos",
             "categoria": "Fertilizante", "unidad_medida": "Galón", "stock_actual": 5}
    res = await validar(_xlsx([FILA_OK, fila2]), db_session)
    assert res.errores == []
    assert res.total_filas == 2 and len(res.filas_ok) == 2

    admin = await _admin(db_session)
    resumen = await importar(db_session, res.filas_ok, admin)
    assert resumen == {"creados": 2}

    prod = (await db_session.execute(
        select(Producto).where(Producto.sku == "IMP-001"))).scalar_one()
    assert prod.stock_actual == Decimal("120")   # via movimiento, no directo
    assert prod.tarifa_iva == Decimal("19")

    mov = (await db_session.execute(
        select(MovimientoInventario).where(MovimientoInventario.producto_id == prod.id))).scalar_one()
    assert mov.tipo.value == "Entrada"
    assert mov.cantidad == Decimal("120")
    assert "Inventario inicial" in (mov.motivo or "")


@pytest.mark.asyncio
async def test_errores_por_fila(db_session):
    # producto existente para provocar el choque de SKU
    db_session.add(Producto(sku="YA-001", nombre="Existente", marca="X",
                            categoria="Otro", unidad_medida="Unidad"))
    await db_session.commit()

    rows = [
        {**FILA_OK, "sku": "IMP-010", "categoria": "Inexistente"},          # enum inválido
        {**FILA_OK, "sku": None, "nombre": None},                           # requeridos faltantes
        {**FILA_OK, "sku": "DUP-1"},                                        # dup en archivo (fila 4)
        {**FILA_OK, "sku": "DUP-1"},                                        # dup en archivo (fila 5)
        {**FILA_OK, "sku": "YA-001"},                                       # ya existe en BD
        {**FILA_OK, "sku": "IMP-011", "precio_venta": "abc"},              # precio no numérico
        {**FILA_OK, "sku": "IMP-012", "tarifa_iva": 12},                   # iva inválido
    ]
    res = await validar(_xlsx(rows), db_session)
    columnas = {e.columna for e in res.errores}
    assert "categoria" in columnas
    assert "sku" in columnas          # faltante + dup + existente
    assert "precio_venta" in columnas
    assert "tarifa_iva" in columnas
    # ninguna de las filas con error entra a filas_ok
    assert all(f.data["sku"] not in {"IMP-010", "YA-001", "IMP-011", "IMP-012"}
               for f in res.filas_ok)


@pytest.mark.asyncio
async def test_encabezado_incompleto_falla(db_session):
    # faltan columnas obligatorias en el encabezado
    headers = ["sku", "nombre"]
    res = await validar(_xlsx([{"sku": "X", "nombre": "Y"}], headers=headers), db_session)
    assert len(res.errores) == 1
    assert "obligatorias" in res.errores[0].mensaje.lower()
    assert res.filas_ok == []


@pytest.mark.asyncio
async def test_filas_vacias_se_ignoran(db_session):
    res = await validar(_xlsx([FILA_OK, {}, {}]), db_session)
    assert res.total_filas == 1 and len(res.filas_ok) == 1


@pytest.mark.asyncio
async def test_generar_plantilla_en_blanco(db_session):
    contenido = generar_plantilla()
    wb = load_workbook(io.BytesIO(contenido))
    assert "Inventario" in wb.sheetnames and "Instrucciones" in wb.sheetnames
    ws = wb["Inventario"]
    headers = [ws.cell(1, i).value for i in range(1, len(EXPECTED_HEADERS) + 1)]
    assert headers == EXPECTED_HEADERS
    assert all(ws.cell(2, i).value is None for i in range(1, len(EXPECTED_HEADERS) + 1))


# ── Endpoints ──────────────────────────────────────────────────
_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_endpoint_plantilla(client, auth_headers):
    resp = await client.get("/api/v1/inventario/plantilla", headers=auth_headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.content[:2] == b"PK"  # xlsx es un zip


@pytest.mark.asyncio
async def test_endpoint_preview_y_commit(client, auth_headers):
    data = _xlsx([FILA_OK])
    r = await client.post(
        "/api/v1/inventario/importar",
        files={"archivo": ("inv.xlsx", data, _MEDIA)}, headers=auth_headers)
    assert r.status_code == 200
    assert r.json()["validas"] == 1 and r.json()["errores"] == []

    r = await client.post(
        "/api/v1/inventario/importar", params={"commit": "true"},
        files={"archivo": ("inv.xlsx", data, _MEDIA)}, headers=auth_headers)
    assert r.status_code == 200 and r.json()["importados"] == 1

    prod = (await client.get(
        "/api/v1/ventas/productos?search=IMP-001", headers=auth_headers)).json()
    assert any(p["sku"] == "IMP-001" for p in prod)


@pytest.mark.asyncio
async def test_endpoint_con_errores_422(client, auth_headers):
    bad = _xlsx([{**FILA_OK, "categoria": "Nope"}])
    r = await client.post(
        "/api/v1/inventario/importar", params={"commit": "true"},
        files={"archivo": ("inv.xlsx", bad, _MEDIA)}, headers=auth_headers)
    assert r.status_code == 422
    assert "errores" in r.json()["detail"]
