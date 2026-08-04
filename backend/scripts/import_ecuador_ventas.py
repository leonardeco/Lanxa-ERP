"""
Importa CUENTAS ECUADOR.xlsx -> tenant Ecuador.
Hoja PAGOS: FECHA, CAMPANA, DETALLE, ASESOR, VALOR VENTA TOTAL, ABONO O PAGO, DEBE

Uso: venv\\Scripts\\python.exe scripts\\import_ecuador_ventas.py <ruta.xlsx> <tenant_id>
"""
import sys
import asyncio
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import date

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
import app.modules.contabilidad.models  # noqa: F401
import app.modules.ventas.models  # noqa: F401
from app.core.database import async_session  # noqa: E402
from app.core.tenancy import set_tenant_id, apply_rls_tenant, reset_tenant_id, tenant_clause  # noqa: E402
from app.modules.ventas.models import Producto, Cliente  # noqa: E402
from app.modules.ventas_diarias.models import (  # noqa: E402
    VentaDiaria, VentaDiariaDetalle, EstadoVentaDiaria,
)
from sqlalchemy import select  # noqa: E402


def _norm(v) -> str:
    return str(v or "").strip().upper()


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _to_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    try:
        from datetime import datetime
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except Exception:
        return None


async def _get_or_create_producto(db, nombre: str, cache: dict) -> int:
    nombre_norm = nombre.strip() or "VENTA"
    if nombre_norm in cache:
        return cache[nombre_norm]
    sku = "EC-" + nombre_norm.upper().replace(" ", "-")[:20]
    existente = await db.scalar(select(Producto).where(Producto.sku == sku, tenant_clause(Producto)))
    if existente:
        cache[nombre_norm] = existente.id
        return existente.id
    p = Producto(sku=sku, nombre=nombre_norm, marca="Super Ozono Ecuador")
    db.add(p)
    await db.flush()
    cache[nombre_norm] = p.id
    return p.id


async def _get_or_create_cliente(db, nombre: str, cache: dict, counter: list) -> int:
    nombre_norm = nombre.strip() or "Ecuador"
    if nombre_norm in cache:
        return cache[nombre_norm]
    counter[0] += 1
    doc = f"EC-{nombre_norm[:15].upper().replace(' ', '-')}-{counter[0]}"
    existente = await db.scalar(select(Cliente).where(Cliente.nit_cc == doc, tenant_clause(Cliente)))
    if existente:
        cache[nombre_norm] = existente.id
        return existente.id
    c = Cliente(nit_cc=doc, razon_social=nombre_norm, tipo_persona="Natural")
    db.add(c)
    await db.flush()
    cache[nombre_norm] = c.id
    return c.id


async def importar(xlsx_path: str, tenant_id: int) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["PAGOS"]

    # Buscar fila de headers
    col_map: dict[str, int] = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=6), start=1):
        if row and any(_norm(c) == "FECHA" for c in row):
            header_row = i
            for idx, cell in enumerate(row):
                n = _norm(cell)
                if n == "FECHA":           col_map["fecha"] = idx
                elif "CAMPANA" in n or "CAMPA" in n: col_map["campana"] = idx
                elif n == "DETALLE":       col_map["detalle"] = idx
                elif n == "ASESOR":        col_map["asesor"] = idx
                elif "VENTA" in n and "TOTAL" in n: col_map["venta"] = idx
                elif "ABONO" in n or "PAGO" in n:   col_map["abono"] = idx
                elif n == "DEBE":          col_map["debe"] = idx
            break

    if not col_map:
        print("No se encontro fila de headers en hoja PAGOS")
        wb.close()
        return

    print(f"Headers en fila {header_row}: {col_map}")

    productos_cache: dict[str, int] = {}
    clientes_cache: dict[str, int] = {}
    counter = [0]
    total_ventas = 0
    omitidas = 0

    async with async_session() as db:
        set_tenant_id(tenant_id)
        await apply_rls_tenant(db, tenant_id)

        for row in ws.iter_rows(values_only=True, min_row=header_row + 1):
            raw = list(row)
            fecha = _to_date(raw[col_map["fecha"]] if "fecha" in col_map else None)
            if fecha is None:
                continue

            campana = str(raw[col_map["campana"]] or "").strip() if "campana" in col_map else ""
            detalle = str(raw[col_map["detalle"]] or "").strip() if "detalle" in col_map else ""
            asesor  = str(raw[col_map["asesor"]]  or "").strip() if "asesor"  in col_map else "N/A"
            venta   = _to_decimal(raw[col_map["venta"]] if "venta" in col_map else None) or Decimal("0")
            abono   = _to_decimal(raw[col_map["abono"]] if "abono" in col_map else None) or Decimal("0")
            debe    = _to_decimal(raw[col_map["debe"]]  if "debe"  in col_map else None) or Decimal("0")

            if venta == 0 and abono == 0:
                omitidas += 1
                continue

            estado = EstadoVentaDiaria.ENTREGADO if debe == 0 else EstadoVentaDiaria.EN_DESTINO

            cliente_nombre = campana or "Ecuador"
            producto_nombre = detalle or campana or "VENTA"

            cliente_id  = await _get_or_create_cliente(db, cliente_nombre, clientes_cache, counter)
            producto_id = await _get_or_create_producto(db, producto_nombre, productos_cache)

            venta_obj = VentaDiaria(
                tenant_id=tenant_id,
                fecha=fecha,
                asesor=asesor or "N/A",
                guia=None,
                cliente_id=cliente_id,
                estado=estado,
                notas=f"Ecuador | Campana: {campana} | Detalle: {detalle}",
            )
            db.add(venta_obj)
            await db.flush()

            detalle_obj = VentaDiariaDetalle(
                venta_diaria_id=venta_obj.id,
                producto_id=producto_id,
                cantidad=Decimal("1"),
                venta=venta,
                abono_1=abono,
                saldo=debe,
            )
            db.add(detalle_obj)
            total_ventas += 1

        await db.commit()

    reset_tenant_id()
    wb.close()
    print(f"Ecuador importado: {total_ventas} ventas | {omitidas} filas omitidas (sin monto).")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: venv\\Scripts\\python.exe scripts\\import_ecuador_ventas.py <ruta.xlsx> <tenant_id>")
        sys.exit(1)
    asyncio.run(importar(sys.argv[1], int(sys.argv[2])))
