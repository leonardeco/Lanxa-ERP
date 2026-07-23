"""
Importa el historico de SUPEROZONO PERU DIARIAS.xlsx (Enero-Julio 2026) al
tenant Peru. Uso unico — no idempotente entre corridas completas (correr
sobre una COPIA de la BD real primero, validar conteos, luego aplicar).

Uso: venv\\Scripts\\python.exe scripts\\import_peru_ventas_diarias.py <ruta.xlsx> <tenant_id>
"""
import sys
import asyncio
from pathlib import Path
from decimal import Decimal, InvalidOperation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.database import async_session  # noqa: E402
from app.core.tenancy import set_tenant_id, apply_rls_tenant, reset_tenant_id  # noqa: E402
from app.modules.ventas.models import Producto, Cliente  # noqa: E402
from app.modules.ventas_diarias.models import (  # noqa: E402
    VentaDiaria, VentaDiariaDetalle, PagoSuelto, EstadoVentaDiaria,
)

HOJAS_MESES = [
    "ENERO SUPEROZONO ", " FEBRERO  SUPEROZONO ", " MARZO  SUPEROZONO  ",
    "ABRIL SUPEROZONO ", "MAYO SUPEROZONO ", "JUNIO SUPEROZONO ",
    "JULIO SUPEROZONO ",
]

_ESTADO_MAP = {
    "ENTREGADO": EstadoVentaDiaria.ENTREGADO,
    "ENTREGDO": EstadoVentaDiaria.ENTREGADO,  # typo real observado en el Excel
    "DEVOLUCION": EstadoVentaDiaria.DEVOLUCION,
    "EN DESTINO": EstadoVentaDiaria.EN_DESTINO,
}


def _norm(v) -> str:
    return str(v or "").strip().upper()


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _encontrar_header(ws) -> tuple[int, dict[str, int]]:
    """Busca la fila de encabezado (contiene 'FECHA') en las primeras 5 filas
    y devuelve (numero_de_fila, {clave_canonica: indice_columna})."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and any(_norm(c) == "FECHA" for c in row):
            mapa: dict[str, int] = {}
            for idx, celda in enumerate(row):
                n = _norm(celda)
                if n == "FECHA":
                    mapa["fecha"] = idx
                elif n == "ASESOR":
                    mapa["asesor"] = idx
                elif n == "GUIA":
                    mapa["guia"] = idx
                elif n == "CODIGO":
                    mapa["codigo"] = idx
                elif n in ("CEDULA", "DNI"):
                    mapa["cedula"] = idx
                elif n == "CLIENTE":
                    mapa["cliente"] = idx
                elif n in ("PEDIDO", "PRODUCTO"):
                    mapa["producto"] = idx
                elif n.startswith("CANT"):
                    mapa["cantidad"] = idx
                elif n in ("VENTA", "V. VENTA", "V VENTA"):
                    mapa["venta"] = idx
                elif n.startswith("RECAUDO 1"):
                    mapa["abono_1"] = idx
                elif n.startswith("RECAUDO 2"):
                    mapa["abono_2"] = idx
                elif n.startswith("RECAUDO"):
                    mapa["abono_1"] = idx
                elif n == "SALDO":
                    mapa["saldo"] = idx
                elif n == "ESTADO":
                    mapa["estado"] = idx
                elif n.startswith("PESOS"):
                    mapa["pesos_c"] = idx
                elif n.startswith("VALOR FLETE"):
                    mapa["valor_flete"] = idx
                elif n.startswith("COMO SE REALIZA"):
                    mapa["forma_pago"] = idx
            return i, mapa
    raise ValueError(f"No se encontro fila de encabezado en hoja '{ws.title}'")


async def _obtener_o_crear_producto(db, nombre: str, cache: dict[str, int]) -> int:
    nombre_norm = nombre.strip()
    if nombre_norm in cache:
        return cache[nombre_norm]
    sku = "PE-" + nombre_norm.upper().replace(" ", "-")[:20]
    existente = await db.scalar(select(Producto).where(Producto.sku == sku))
    if existente:
        cache[nombre_norm] = existente.id
        return existente.id
    producto = Producto(sku=sku, nombre=nombre_norm, marca="Super Ozono Peru")
    db.add(producto)
    await db.flush()
    cache[nombre_norm] = producto.id
    return producto.id


async def _obtener_o_crear_cliente(
    db, documento: str | None, nombre: str, cache: dict[str, int], contador: list[int]
) -> int:
    if not documento:
        contador[0] += 1
        documento = f"SIN-DOC-{contador[0]}"
    if documento in cache:
        return cache[documento]
    existente = await db.scalar(select(Cliente).where(Cliente.nit_cc == documento))
    if existente:
        cache[documento] = existente.id
        return existente.id
    cliente = Cliente(
        nit_cc=documento, razon_social=nombre.strip() or documento,
        tipo_persona="Natural",
    )
    db.add(cliente)
    await db.flush()
    cache[documento] = cliente.id
    return cliente.id


async def importar(xlsx_path: str, tenant_id: int) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    async with async_session() as db:
        set_tenant_id(tenant_id)
        await apply_rls_tenant(db, tenant_id)

        productos_cache: dict[str, int] = {}
        clientes_cache: dict[str, int] = {}
        contador_sin_doc = [0]
        total_ventas = 0
        total_pagos_sueltos = 0

        for nombre_hoja in HOJAS_MESES:
            if nombre_hoja not in wb.sheetnames:
                print(f"AVISO: hoja '{nombre_hoja}' no encontrada, se omite")
                continue
            ws = wb[nombre_hoja]
            header_row, col = _encontrar_header(ws)

            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
                if not row or row[col["fecha"]] is None:
                    continue

                fecha = row[col["fecha"]]
                if hasattr(fecha, "date"):
                    fecha = fecha.date()

                cliente_nombre = row[col.get("cliente", -1)] if "cliente" in col else None
                producto_nombre = row[col.get("producto", -1)] if "producto" in col else None

                # Fila de "pago suelto": sin producto ni guia, cliente suele
                # empezar con "PAGO " o la columna ESTADO dice literalmente
                # "PAGO" (algunos meses usan CLIENTE="YAPE"/nombre de persona
                # en vez del prefijo "PAGO ") — se importa como PagoSuelto,
                # no como venta.
                estado_fila_raw = _norm(row[col["estado"]]) if "estado" in col else ""
                es_pago_suelto = (
                    cliente_nombre and "PAGO" in _norm(cliente_nombre)
                ) or estado_fila_raw == "PAGO"
                if not producto_nombre and es_pago_suelto:
                    abono = _to_decimal(row[col["abono_1"]]) if "abono_1" in col else None
                    if abono:
                        db.add(PagoSuelto(
                            fecha=fecha,
                            cliente_texto=str(cliente_nombre).strip(),
                            monto=abono,
                            revisado=False,
                            notas="Importado de Excel — sin vinculo confirmado a una venta.",
                        ))
                        total_pagos_sueltos += 1
                    continue

                if not producto_nombre:
                    continue  # fila vacia / de total, sin producto real

                cliente_id = await _obtener_o_crear_cliente(
                    db,
                    str(row[col["cedula"]]).strip() if "cedula" in col and row[col["cedula"]] else None,
                    str(cliente_nombre or "Sin nombre"),
                    clientes_cache, contador_sin_doc,
                )
                producto_id = await _obtener_o_crear_producto(
                    db, str(producto_nombre), productos_cache)

                estado_raw = _norm(row[col["estado"]]) if "estado" in col else ""
                estado = _ESTADO_MAP.get(estado_raw, EstadoVentaDiaria.PENDIENTE)

                venta = VentaDiaria(
                    fecha=fecha,
                    asesor=str(row[col["asesor"]]).strip() if "asesor" in col and row[col["asesor"]] else None,
                    guia=str(row[col["guia"]]).strip() if "guia" in col and row[col["guia"]] else None,
                    codigo_guia=str(row[col["codigo"]]).strip() if "codigo" in col and row[col["codigo"]] else None,
                    cliente_id=cliente_id,
                    estado=estado,
                    forma_pago=str(row[col["forma_pago"]]).strip() if "forma_pago" in col and row[col["forma_pago"]] else None,
                    notas=f"Importado de hoja '{nombre_hoja}'",
                )
                db.add(venta)
                await db.flush()

                venta_val = _to_decimal(row[col["venta"]]) if "venta" in col else None
                abono_1 = _to_decimal(row[col["abono_1"]]) if "abono_1" in col else None
                abono_2 = _to_decimal(row[col["abono_2"]]) if "abono_2" in col else None
                saldo = (venta_val or Decimal("0")) - (abono_1 or Decimal("0")) - (abono_2 or Decimal("0"))

                db.add(VentaDiariaDetalle(
                    venta_diaria_id=venta.id,
                    producto_id=producto_id,
                    cantidad=_to_decimal(row[col["cantidad"]]) or Decimal("1"),
                    venta=venta_val,
                    abono_1=abono_1,
                    abono_2=abono_2,
                    saldo=saldo,
                    pesos_c=_to_decimal(row[col["pesos_c"]]) if "pesos_c" in col else None,
                    valor_flete=_to_decimal(row[col["valor_flete"]]) if "valor_flete" in col else None,
                ))
                total_ventas += 1

            print(f"Hoja '{nombre_hoja}': procesada.")

        await db.commit()
        reset_tenant_id()

    print(f"Importacion completa: {total_ventas} ventas diarias, {total_pagos_sueltos} pagos sueltos.")
    print("Revisar 'PESOS C' / 'VALOR FLETE' y los pagos sueltos con la auxiliar de Peru antes de usarlos en reportes.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Uso: import_peru_ventas_diarias.py <ruta.xlsx> <tenant_id>")
    asyncio.run(importar(sys.argv[1], int(sys.argv[2])))
