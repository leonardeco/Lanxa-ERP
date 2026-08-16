"""
Importación de inventario inicial desde Excel (.xlsx) — #2 (parcial).

Fuente única de columnas (EXPECTED_HEADERS). Genera la plantilla en blanco,
valida el archivo fila por fila (sin escribir nada) y, si no hay errores, crea
los productos con su stock inicial registrado en el kardex. Todo o nada.

El asiento de apertura contable queda FUERA de alcance (depende del método de
costeo, pendiente de la contadora).
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.ventas.models import Producto, CategoriaProducto, UnidadMedida
from app.modules.contabilidad.models import CentroCosto
from app.modules.inventario.models import TipoMovimientoInventario, OrigenMovimiento
from app.modules.inventario.service import registrar_movimiento
from app.modules.inventario.lotes import entrada_lote
from app.modules.auditoria.service import registrar_auditoria

REQUERIDAS = [
    "sku", "nombre", "categoria", "marca", "unidad_medida",
    "precio_venta", "tarifa_iva", "stock_actual",
]
OPCIONALES = [
    "contenido_neto", "precio_costo", "stock_minimo",
    "registro_ica", "centro_costo_codigo", "descripcion", "notas",
    "codigo_lote", "fecha_vencimiento",
]
EXPECTED_HEADERS = REQUERIDAS + OPCIONALES

_CATEGORIAS = {c.value for c in CategoriaProducto}
_UNIDADES = {u.value for u in UnidadMedida}
_IVA_VALIDAS = {Decimal("19"), Decimal("5"), Decimal("0")}
_HOJA = "Inventario"


@dataclass
class ErrorFila:
    fila: int
    columna: str
    mensaje: str


@dataclass
class FilaProducto:
    fila: int
    data: dict
    stock_inicial: Decimal
    costo: Decimal | None
    codigo_lote: str | None = None
    fecha_vencimiento: date | None = None


@dataclass
class ResultadoValidacion:
    filas_ok: list[FilaProducto] = field(default_factory=list)
    errores: list[ErrorFila] = field(default_factory=list)
    total_filas: int = 0


def _s(v) -> str | None:
    """Celda de texto → string recortado o None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> Decimal:
    """Celda numérica → Decimal. Lanza ValueError si no es un número."""
    if v is None or isinstance(v, bool):
        raise ValueError("vacío o no numérico")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip()
    if not s:
        raise ValueError("vacío")
    if "," in s and "." not in s:  # tolera coma decimal colombiana
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError("no numérico")


def _fecha(v) -> date:
    """Celda de fecha → date. Acepta datetime/date de Excel o texto ISO (YYYY-MM-DD)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        raise ValueError("vacío")
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        raise ValueError("fecha inválida (usa AAAA-MM-DD)")


async def validar(contenido: bytes, db: AsyncSession) -> ResultadoValidacion:
    """Lee y valida el .xlsx. No escribe nada en la BD."""
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:  # archivo no es un xlsx válido
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}") from exc

    ws = wb[_HOJA] if _HOJA in wb.sheetnames else wb.active
    if ws is None:
        raise ValueError("El archivo no tiene una hoja de datos.")
    res = ResultadoValidacion()

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        res.errores.append(ErrorFila(1, "-", "El archivo está vacío."))
        return res

    encabezados = [(_s(c) or "").lower() for c in filas[0]]
    col = {name: encabezados.index(name) for name in EXPECTED_HEADERS if name in encabezados}
    faltantes = [name for name in REQUERIDAS if name not in col]
    if faltantes:
        res.errores.append(ErrorFila(1, ", ".join(faltantes),
                                     "Faltan columnas obligatorias en el encabezado."))
        return res

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    existentes = set((await db.execute(select(Producto.sku))).scalars().all())
    cc_map = {
        codigo: cid
        for codigo, cid in (await db.execute(select(CentroCosto.codigo, CentroCosto.id))).all()
    }
    skus_archivo: dict[str, int] = {}

    for i, row in enumerate(filas[1:], start=2):
        # saltar filas totalmente vacías
        if all(get(row, n) in (None, "") for n in EXPECTED_HEADERS):
            continue
        res.total_filas += 1
        errores_fila: list[ErrorFila] = []

        sku = _s(get(row, "sku"))
        nombre = _s(get(row, "nombre"))
        marca = _s(get(row, "marca"))
        categoria = _s(get(row, "categoria"))
        unidad = _s(get(row, "unidad_medida"))

        if not sku:
            errores_fila.append(ErrorFila(i, "sku", "El SKU es obligatorio."))
        else:
            if sku in skus_archivo:
                errores_fila.append(ErrorFila(
                    i, "sku",
                    f"SKU '{sku}' repetido (ya está en la fila {skus_archivo[sku]})."))
            elif sku in existentes:
                errores_fila.append(ErrorFila(
                    i, "sku",
                    f"El SKU '{sku}' ya existe en el sistema; no se sobreescribe."))
            skus_archivo.setdefault(sku, i)

        if not nombre:
            errores_fila.append(ErrorFila(i, "nombre", "El nombre es obligatorio."))
        if not marca:
            errores_fila.append(ErrorFila(i, "marca", "La marca es obligatoria."))
        if not categoria:
            errores_fila.append(ErrorFila(i, "categoria", "La categoría es obligatoria."))
        elif categoria not in _CATEGORIAS:
            errores_fila.append(ErrorFila(
                i, "categoria",
                f"Categoría inválida '{categoria}'. Válidas: {', '.join(sorted(_CATEGORIAS))}."))
        if not unidad:
            errores_fila.append(ErrorFila(i, "unidad_medida", "La unidad es obligatoria."))
        elif unidad not in _UNIDADES:
            errores_fila.append(ErrorFila(
                i, "unidad_medida",
                f"Unidad inválida '{unidad}'. Válidas: {', '.join(sorted(_UNIDADES))}."))

        def _positivo(nombre_col, obligatorio):
            crudo = get(row, nombre_col)
            if crudo in (None, ""):
                if obligatorio:
                    errores_fila.append(ErrorFila(i, nombre_col, "Valor obligatorio."))
                return None
            try:
                val = _num(crudo)
            except ValueError:
                errores_fila.append(ErrorFila(i, nombre_col, f"'{crudo}' no es un número válido."))
                return None
            if val < 0:
                errores_fila.append(ErrorFila(i, nombre_col, "No puede ser negativo."))
                return None
            return val

        precio_venta = _positivo("precio_venta", True)
        stock = _positivo("stock_actual", True)
        precio_costo = _positivo("precio_costo", False)

        iva = None
        iva_crudo = get(row, "tarifa_iva")
        if iva_crudo in (None, ""):
            errores_fila.append(ErrorFila(i, "tarifa_iva", "El IVA es obligatorio (19, 5 o 0)."))
        else:
            try:
                iva = _num(iva_crudo)
                if iva not in _IVA_VALIDAS:
                    errores_fila.append(ErrorFila(i, "tarifa_iva", "El IVA debe ser 19, 5 o 0."))
                    iva = None
            except ValueError:
                errores_fila.append(ErrorFila(i, "tarifa_iva", f"'{iva_crudo}' no es un IVA válido."))

        stock_min = None
        sm_crudo = get(row, "stock_minimo")
        if sm_crudo not in (None, ""):
            try:
                stock_min = int(_num(sm_crudo))
                if stock_min < 0:
                    errores_fila.append(ErrorFila(i, "stock_minimo", "No puede ser negativo."))
                    stock_min = None
            except ValueError:
                errores_fila.append(ErrorFila(i, "stock_minimo", "Debe ser un entero."))

        centro_costo_id = None
        cc_codigo = _s(get(row, "centro_costo_codigo"))
        if cc_codigo:
            centro_costo_id = cc_map.get(cc_codigo)
            if centro_costo_id is None:
                errores_fila.append(ErrorFila(
                    i, "centro_costo_codigo",
                    f"No existe un centro de costo con código '{cc_codigo}'."))

        # Lote + vencimiento (opt-in): si viene código, el producto queda con
        # control de lote y su stock inicial entra como un Lote.
        codigo_lote = _s(get(row, "codigo_lote"))
        fecha_venc = None
        fv_crudo = get(row, "fecha_vencimiento")
        if fv_crudo not in (None, ""):
            try:
                fecha_venc = _fecha(fv_crudo)
            except ValueError as exc:
                errores_fila.append(ErrorFila(i, "fecha_vencimiento", str(exc)))
        if fecha_venc is not None and not codigo_lote:
            errores_fila.append(ErrorFila(
                i, "codigo_lote",
                "Hay fecha de vencimiento pero falta el código de lote."))
        if codigo_lote and (stock is None or stock <= 0):
            errores_fila.append(ErrorFila(
                i, "stock_actual",
                "Un producto con lote necesita stock inicial mayor a cero."))

        if errores_fila:
            res.errores.extend(errores_fila)
            continue

        data = {
            "sku": sku,
            "nombre": nombre,
            "marca": marca,
            "categoria": CategoriaProducto(categoria),
            "unidad_medida": UnidadMedida(unidad),
            "precio_venta": precio_venta,
            "tarifa_iva": iva,
            "descripcion": _s(get(row, "descripcion")),
            "contenido_neto": _s(get(row, "contenido_neto")),
            "registro_ica": _s(get(row, "registro_ica")),
            "notas": _s(get(row, "notas")),
        }
        if precio_costo is not None:
            data["precio_costo"] = precio_costo
        if stock_min is not None:
            data["stock_minimo"] = stock_min
        if centro_costo_id is not None:
            data["centro_costo_id"] = centro_costo_id
        if codigo_lote:
            data["controla_lote"] = True

        res.filas_ok.append(FilaProducto(
            fila=i, data=data, stock_inicial=stock, costo=precio_costo,
            codigo_lote=codigo_lote, fecha_vencimiento=fecha_venc))

    return res


async def importar(db: AsyncSession, filas_ok: list[FilaProducto], usuario) -> dict:
    """Crea los productos + su entrada de stock inicial en el kardex. Atómico."""
    creados = 0
    for f in filas_ok:
        producto = Producto(**f.data, stock_actual=Decimal("0"))
        db.add(producto)
        await db.flush()
        if f.stock_inicial and f.stock_inicial > 0:
            if f.codigo_lote:
                await entrada_lote(
                    db,
                    producto_id=producto.id,
                    cantidad=f.stock_inicial,
                    codigo_lote=f.codigo_lote,
                    fecha_vencimiento=f.fecha_vencimiento,
                    origen=OrigenMovimiento.AJUSTE_MANUAL,
                    costo_unitario=f.costo,
                    usuario_id=usuario.id if usuario else None,
                    motivo="Inventario inicial (importación)",
                )
            else:
                await registrar_movimiento(
                    db,
                    producto_id=producto.id,
                    tipo=TipoMovimientoInventario.ENTRADA,
                    origen=OrigenMovimiento.AJUSTE_MANUAL,
                    cantidad=f.stock_inicial,
                    motivo="Inventario inicial (importación)",
                    usuario_id=usuario.id if usuario else None,
                    costo_unitario=f.costo,
                )
        registrar_auditoria(
            db, usuario, "Crear", "Producto", producto.id,
            f"Producto {producto.sku} — {producto.nombre} (importación inventario inicial)",
        )
        creados += 1
    await db.commit()
    return {"creados": creados}


def generar_plantilla() -> bytes:
    """Genera el .xlsx en blanco (hoja Inventario con desplegables + Instrucciones)."""
    req_fill = PatternFill("solid", fgColor="1F7A3D")
    opt_fill = PatternFill("solid", fgColor="6B7280")
    hdr_font = Font(bold=True, color="FFFFFF")

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # hoja por defecto de un Workbook nuevo
    ws.title = _HOJA
    maxrow = 1000

    anchos = {"sku": 14, "nombre": 36, "descripcion": 30,
              "codigo_lote": 16, "fecha_vencimiento": 16}
    numfmt = {"precio_venta": "#,##0.00", "precio_costo": "#,##0.00",
              "tarifa_iva": "0", "stock_actual": "#,##0.###", "stock_minimo": "0",
              "fecha_vencimiento": "yyyy-mm-dd"}
    for idx, name in enumerate(EXPECTED_HEADERS, start=1):
        letter = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = hdr_font
        cell.fill = req_fill if name in REQUERIDAS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[letter].width = anchos.get(name, 15)
        etiqueta = "OBLIGATORIO." if name in REQUERIDAS else "Opcional."
        cell.comment = Comment(etiqueta, "Plantilla")
        if name in numfmt:
            for r in range(2, maxrow + 1):
                ws.cell(row=r, column=idx).number_format = numfmt[name]

    validaciones = {
        "categoria": '"' + ",".join(sorted(_CATEGORIAS)) + '"',
        "unidad_medida": '"Litro,Galón,Kilo,Unidad,Caja,Caneca"',
        "tarifa_iva": '"19,5,0"',
    }
    for name, formula in validaciones.items():
        letter = get_column_letter(EXPECTED_HEADERS.index(name) + 1)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.errorTitle = "Valor inválido"
        dv.error = "Elige un valor de la lista."
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{maxrow}")

    ws.freeze_panes = "A2"

    ins = wb.create_sheet("Instrucciones")
    guia = [
        "Plantilla de inventario inicial — Lanxa ERP",
        "",
        "1. Una fila por producto en la hoja 'Inventario'.",
        "2. Encabezado VERDE = obligatorio; GRIS = opcional.",
        "3. categoria, unidad_medida e IVA se eligen del desplegable.",
        "4. Precios y stock: punto decimal, sin separador de miles (ej: 85000.00).",
        "5. El SKU no se puede repetir.",
        "6. codigo_lote + fecha_vencimiento: solo para productos con control de lote "
        "(perecederos/orgánicos). Si pones código, el stock inicial entra como ese "
        "lote y el producto queda con control de lote. Deja ambos en blanco para los "
        "durables.",
        "7. fecha_vencimiento en formato AAAA-MM-DD (ej: 2027-06-30).",
        "",
        "categoria válidas: " + ", ".join(sorted(_CATEGORIAS)),
        "unidad_medida válidas: Litro, Galón, Kilo, Unidad, Caja, Caneca",
        "tarifa_iva: 19, 5 o 0",
    ]
    for r, linea in enumerate(guia, start=1):
        ins.cell(row=r, column=1, value=linea)
    ins.column_dimensions["A"].width = 70

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
